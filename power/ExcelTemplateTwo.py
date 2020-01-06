# -*- coding:utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
import time
# 4个数据集合 传进封装类 实体

class Excel_TemplateTwo(object):
    def __init__(self, board_voltage, mu_voltage, board_current, mu_current, board_num):
        self.board_voltage_ = board_voltage
        self.mu_voltage_ = mu_voltage
        self.board_current_ = board_current
        self.mu_current_ = mu_current
        self.board_num_ = board_num
        self.setup_excel()

    def setup_excel(self):
        """创建excel模版"""
        wb = openpyxl.Workbook()
        # 如果不给sheet名称则直接创建默认的sheet1名称
        if self.board_num_ is None:
            ws = wb.create_sheet('sheet1', 0)
        else:
            ws = wb.create_sheet(str(self.board_num_), 0)
        # 居中格式设置
        ali = Alignment(horizontal='center', vertical='center')
        # 边框设置
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        myborder = Border(left=left, right=right, top=top, bottom=bottom)

        # 5种颜色配置
        SpringGreen2 = PatternFill("solid", fgColor="6A5ACD")
        DeepSkyBlue = PatternFill("solid", fgColor="00BFFF")
        Turquoise = PatternFill("solid", fgColor="40E0D0")
        Yellow3 = PatternFill("solid", fgColor="CDCD00")
        PaleGoldenrod = PatternFill("solid", fgColor="EEE8AA")
        white = PatternFill("solid", fgColor="FFFFFF")

        # 设置 1 2 3 4 5 6 列的宽度
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20.0
        # 设置小电流模式
        for t in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            ws[t].alignment = ali
            ws[t].border = myborder
            ws[t].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)

        ws['A1'].value = '大电流模式'
        ws['B1'].value = '机器编码'

        ws.merge_cells('B1:C1')
        ws['D1'].value = '板卡条码'+self.board_num_
        ws.merge_cells('D1:F1')

        ws['A2'].value = '设置电压'
        ws['B2'].value = '大电流模式采样次数'
        ws['C2'].value = '8001电压读数（mV）'
        ws['D2'].value = '万用表电压读数（mV）'
        ws['E2'].value = '8001电流读数(mA)'
        ws['F2'].value = '万用表电流读数(mA)'

        for t in ['A2', 'B2', 'C2', 'D2', 'E2', 'F2']:
            ws[t].alignment = ali
            ws[t].border = myborder
            ws[t].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)

        # 第一轮老化测试
        ws['A3'].value = '第一轮老化测试'
        ws['A3'].alignment = ali
        ws['A3'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in ['A3', 'B3', 'C3', 'D3', 'E3', 'F3']:
            ws[t].border = myborder
            ws[t].fill = white
        ws.merge_cells('A3:F3')

        ws['A4'].value = '3.6V'
        ws['A4'].alignment = ali
        ws['A4'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (4, 9):
            ws['A'+str(t)].border = myborder
            ws['A' + str(t)].fill = SpringGreen2
            ws['B' + str(t) ].fill = SpringGreen2
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A4:A8')

        ws['A9'].value = '3.8V'
        ws['A9'].alignment = ali
        ws['A9'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (9, 14):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A9:A13')

        ws['A14'].value = '4.0V'
        ws['A14'].alignment = ali
        ws['A14'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (14, 19):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Turquoise
            ws['B' + str(t)].fill = Turquoise
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A14:A18')

        ws['A19'].value = '4.1V'
        ws['A19'].alignment = ali
        ws['A19'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (19, 24):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Yellow3
            ws['B' + str(t)].fill = Yellow3
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A19:A23')

        ws['A24'].value = '4.2V'
        ws['A24'].alignment = ali
        ws['A24'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (24, 29):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A24:A28')

        # 第二轮老化测试
        ws['A29'].value = '第二轮老化测试'
        ws['A29'].alignment = ali
        ws['A29'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in ['A29', 'B29', 'C29', 'D29', 'E29', 'F29']:
            ws[t].border = myborder
            ws[t].fill = white
        ws.merge_cells('A29:F29')

        ws['A30'].value = '3.6V'
        ws['A30'].alignment = ali
        ws['A30'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (30, 35):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = SpringGreen2
            ws['B'+ str(t)].fill = SpringGreen2
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A30:A34')

        ws['A35'].value = '3.8V'
        ws['A35'].alignment = ali
        ws['A35'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (35, 40):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A35:A39')

        ws['A40'].value = '4.0V'
        ws['A40'].alignment = ali
        ws['A40'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (40, 45):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Turquoise
            ws['B' + str(t)].fill = Turquoise
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A40:A44')

        ws['A45'].value = '4.1V'
        ws['A45'].alignment = ali
        ws['A45'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (45, 50):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Yellow3
            ws['B' + str(t)].fill = Yellow3
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A45:A49')

        ws['A50'].value = '4.2V'
        ws['A50'].alignment = ali
        ws['A50'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (50, 55):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A50:A54')

        # 第三轮老化测试
        ws['A55'].value = '第三轮老化测试'
        ws['A55'].alignment = ali
        ws['A55'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in ['A55', 'B55', 'C55', 'D55', 'E55', 'F55']:
            ws[t].border = myborder
            ws[t].fill = white
        ws.merge_cells('A55:F55')

        ws['A56'].value = '3.6V'
        ws['A56'].alignment = ali
        ws['A56'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (56, 61):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = SpringGreen2
            ws['B' + str(t)].fill = SpringGreen2
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A56:A60')

        ws['A61'].value = '3.8V'
        ws['A61'].alignment = ali
        ws['A61'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (61, 66):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A61:A65')

        ws['A66'].value = '4.0V'
        ws['A66'].alignment = ali
        ws['A66'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (66, 71):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Turquoise
            ws['B' + str(t)].fill = Turquoise
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A66:A70')

        ws['A71'].value = '4.1V'
        ws['A71'].alignment = ali
        ws['A71'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (71, 76):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Yellow3
            ws['B' + str(t)].fill = Yellow3
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A71:A75')

        ws['A76'].value = '4.2V'
        ws['A76'].alignment = ali
        ws['A76'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (76, 80):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A76:A80')

        q = [q for q in range(1, 6)]
        color_array = [SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,
                       DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,
                       Turquoise,Turquoise,Turquoise,Turquoise,Turquoise,
                       Yellow3,Yellow3,Yellow3,Yellow3,Yellow3,
                       PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod]

        for i,j,h in zip(range(4, 29), q*5, color_array):
            ws.cell(i, 2, j)
            ws.cell(i,2).fill = h
            ws.cell(i,2).border = myborder

        for i, j,h in zip(range(30, 55), q * 5, color_array):
            ws.cell(i, 2, j)
            ws.cell(i, 2).fill = h
            ws.cell(i, 2).border = myborder

        for i, j,h in zip(range(56, 81), q * 5, color_array):
            ws.cell(i, 2, j)
            ws.cell(i, 2).fill = h
            ws.cell(i, 2).border = myborder

        # 从这里开始开始数据添加
        board_voltage_one = self.board_voltage_[0:25]
        board_voltage_two = self.board_voltage_[25:50]
        board_voltage_three = self.board_voltage_[50:]

        board_current_one = self.board_current_[0:25]
        board_current_two = self.board_current_[25:50]
        board_current_three = self.board_current_[50:]

        mu_voltage_one = self.mu_voltage_[0:25]
        mu_voltage_two = self.mu_voltage_[25:50]
        mu_voltage_three = self.mu_voltage_[50:]

        mu_current_one = self.mu_current_[0:25]
        mu_current_two = self.mu_current_[25:50]
        mu_current_three = self.mu_current_[50:]

        # 第一组数据添加
        for j,i in zip(range(4, 29), board_voltage_one):
            ws.cell(j, 3, i)

        for j,i in zip(range(4, 29),mu_voltage_one):
            ws.cell(j, 4, i)

        for j, i in zip(range(4, 29), board_current_one):
            ws.cell(j, 5, i)

        for j, i in zip(range(4, 29), mu_current_one):
            ws.cell(j, 6, i)

        # 第二组数据添加
        for j,i in zip(range(30, 55), board_voltage_two):
            ws.cell(j, 3, i)

        for j,i in zip(range(30, 55),mu_voltage_two):
            ws.cell(j, 4, i)

        for j, i in zip(range(30, 55), board_current_two):
            ws.cell(j, 5, i)

        for j, i in zip(range(30, 55), mu_current_two):
            ws.cell(j, 6, i)


        # 第三组数据添加
        for j,i in zip(range(56, 81), board_voltage_three):
            ws.cell(j, 3, i)

        for j,i in zip(range(56, 81), mu_voltage_three):
            ws.cell(j, 4, i)

        for j, i in zip(range(56, 81), board_current_three):
            ws.cell(j, 5, i)

        for j, i in zip(range(56, 81), mu_current_three):
            ws.cell(j, 6, i)

        prev_msg = str(time.localtime().tm_year) + '_' + str(time.localtime().tm_mon) + '_' + str(
            time.localtime().tm_mday) + '_' + str(time.localtime().tm_hour) + '_' + str(
            time.localtime().tm_min) + '_' + str(
            time.localtime().tm_sec)
        if self.board_num_ is None:
            middle_msg = 'sheet1'
        else:
            middle_msg = str(self.board_num_)
        next_msg = '100R'
        msg = prev_msg+'_'+middle_msg+'_'+next_msg
        wb.save('./'+msg+'.xlsx')


def main():
    alist = [3594, 3575, 3574, 3586, 3608, 3791, 3789, 3781, 3839, 3823, 3978, 3974, 4000, 4014, 3979, 4089, 4087, 4106, 4064, 4089, 4211, 4193, 4202, 4208, 4187, 3611, 3595, 3635, 3588, 3571, 3810, 3790, 3836, 3771, 3797, 3999, 3968, 3985, 3992, 4003, 4094, 4059, 4098, 4092, 4074, 4199, 4175, 4161, 4188, 4175, 3594, 3594, 3589, 3573, 3585, 3800, 3777, 3784, 3789, 3823, 3975, 3970, 4001, 3975, 4015, 4096, 4104, 4104, 4057, 4116, 4189, 4198, 4173, 4193, 4202]
    blist = [3597.59689, 3566.86626, 3582.57248, 3593.38684, 3599.84626, 3745.62806, 3787.17431, 3777.36686, 3827.28425, 3825.79079, 3915.57689, 3971.36577, 3986.08519, 4001.10843, 3955.4602, 4125.05681, 4081.69956, 4080.1682, 4097.87961, 4087.91354, 4185.56944, 4167.33518, 4170.94763, 4189.68034, 4172.03757, 3558.72537, 3606.21799, 3572.64609, 3581.08564, 3600.6408, 3793.39601, 3773.00673, 3785.73324, 3757.48383, 3809.75624, 3983.3845, 3963.18079, 3974.56226, 4030.31671, 3992.67161, 4078.78327, 4043.58401, 4065.80892, 4057.08027, 4064.53497, 4225.2815, 4174.54702, 4157.04808, 4179.23612, 4159.57662, 3559.67729, 3605.34009, 3565.82298, 3577.13703, 3562.14356, 3840.17373, 3783.0411, 3778.49626, 3780.6824, 3784.10816, 3977.5952, 3960.15632, 3991.22436, 3992.67057, 3998.45225, 4093.45185, 4077.66654, 4077.43191, 4058.70541, 4099.47178, 4233.05874, 4183.02799, 4158.1451, 4173.95594, 4182.52741]
    clist = [105.7, 173.7, 273.7, 649.7, 1428.2, 0.2, 184, 289, 696.7, 1511.2, 0.3, 192.7, 306.7, 727.5, 1585.1, 120.8, 198.1, 315.1, 738.6, 1621.6, 124.6, 203.3, 321.9, 759, 1663.3, 0, 174.5, 277.8, 649.9, 1422 , 112.4, 183.9, 294.1, 686.5, 1501.3, 0.3, 193.7, 306.2, 721.2, 1581.4, 121.1, 198.1, 315.5, 743.2, 1618.3, 124.2, 202.9, 320.7, 768.1, 1671.6, 0, 174.4, 279.4, 645.3, 1422.5, 112, 183.3, 289.9, 683.2, 1510 , 0.5, 192.8, 307, 720.6, 1588.8, 0.5, 129.3, 313.4, 730, 1615.7, 123.9, 203.6, 319.7, 751.2, 1663.5]
    dlist = [107.260793, 176.670212, 276.53507, 651.983634, 1433.8283, 0.01091728, 185.582212, 292.755282, 698.061243, 1519.67552, 0.011730109, 195.936078, 309.723241, 729.007415, 1591.9457, 122.018937, 200.822063, 317.79457, 742.338394, 1626.13994, 125.499833, 205.971196, 324.940867, 763.797198, 1670.77956, 0.010235908, 175.735711, 280.838644, 652.577347, 1425.20922, 113.508903, 186.308558, 296.459948, 689.037551, 1508.48452, 0.011716185, 195.444709, 308.055494, 725.037578, 1588.80405, 122.218566, 200.571705, 317.405109, 747.499879, 1626.2085, 125.172164, 205.416831, 323.626179, 763.729914, 1666.8969, 0.010164204, 175.732899, 277.703695, 649.265954, 1430.61626, 113.145348, 185.710018, 292.855559, 687.889821, 1516.54431, 0.011624727, 195.413707, 309.77064, 722.80724, 1593.21382, 0.012071344, 130.117927, 317.807833, 734.281924, 1620.9838, 124.986929, 206.250978, 321.48137, 753.671331, 1670.4512]
    print(len(alist))
    print(len(blist))
    print(len(clist))
    print(len(dlist))
    e = Excel_TemplateTwo(alist,blist, clist, dlist, '1')
    # print((1 and 1 and 1 and 1 and 1) is False)

if __name__ == '__main__':
    main()